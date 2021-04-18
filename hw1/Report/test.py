import openpyxl
import statistics
from openpyxl.workbook import Workbook
from openpyxl.styles import Font,NamedStyle


header = ['Name', 'Email', 'Mobile', 'Current location']
new_data = [['name1', 'email1@yahoo.com', 9929283421.0, 'xxxx'],
            ['name2', 'email2@xyz.com', 9994191988.0, 'xxxx']]

wb = Workbook()

dest_filename = 'test_book.xlsx'

ws1 = wb.active
header_style = NamedStyle(name="header_style")
header_style.font = Font(bold=True)
ws1.append(header)
header_row = ws1[1]
for cell in header_row:
    cell.style = header_style

for row in new_data:
    ws1.append(row)


wb.save(filename = dest_filename)

"""
workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello_world.xlsx")


wb_obj = openpyxl.load_workbook("Grades.xlsx", data_only=True)
sheet = wb_obj.active
values = sheet["A"]
countries = sheet["K"]
my_list = []
c_list = []
reproduction_rate = sheet["M"]

my_set = set()
rate_list = []
i = 0
for element in countries[1:]:
    c_list.append(element.value)
for element in reproduction_rate[1:]:
    my_list.append(element.value)

for country in c_list:
    size = len(my_set)
    my_set.add(country)
    if len(my_set) == size or i == 0:
        if my_list[i] is not None:
            rate_list.append(my_list[i])
    else:
        print(c_list[i-1])
        print(rate_list)
        print("max:", min([5]))
        if my_list[i] is not None:
            rate_list = [my_list[i]]
        else:
            rate_list = []
    i = i + 1


for element in values[1:]:
    my_list.append(element.value)
print(my_list)

indexes = []
i = 1
for element in values:
    size = len(myset)
    myset.add(element.value)
    if len(myset) != size:
        indexes.append(i - 1)
    i = i + 1

# print(indexes[2:])
for index in indexes[2:]:
    print(values[index - 1].value, " ", grades[index - 1].value)


mylist = []
for user in values:
    mylist.append(user.value)
# print(mylist)
min_date = min(mylist)
print(min_date)
indexes = [i for i, x in enumerate(mylist) if x == min_date]
countries = sheet["B"]
for index in indexes:
    print(countries[index].value)
print(indexes)

for user in values:
    myset.add(user.value)
print("Normal Set:")
print(myset)
print("Sorted:")
last = sorted(myset, key=lambda d: tuple(map(int, d.split('-'))))
print(last)
"""
